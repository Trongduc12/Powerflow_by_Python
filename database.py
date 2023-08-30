#-------------------------------------------------------------------------------
# Name:        database
# Purpose:
#
# Author:      Trong duc
#
# Created:     27/07/2023
# Copyright:   (c) Trong 2023
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import sqlite3
import openpyxl
import sys
import codecs

# translate Vietnamese
sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

class database():
    def __init__(self, excel_file, db_file):
        wb = openpyxl.load_workbook(excel_file)
        self.Tab_Col_Excel = self.__Tab_Col_Excel__(wb)

        self.conn = sqlite3.connect(db_file)
        self.cursor = self.conn.cursor()

        self.Tab_Col_SQL, self.Len_Tab = self.__Tab_Col_SQL__()

        self.Insert, self.Update = self.__getDataExcel__(wb)
    def __getDataExcel__(self,wb):

        # List Insert
        res={}

        # List Update
        res1={}

        """ Tạo dict Insert và dict Update """
        # Bảng trong SQL trùng với sheet Excel
        for k in self.Tab_Col_SQL:
            if k in self.Tab_Col_Excel.keys():
                sht= wb[k]

                # Tạo list data Insert
                list1= []

                # Lấy dữ liệu từng hàng trong sheet
                for row in sht.iter_rows(min_row=3, values_only=True):

                    # Lấy hết data 1 hàng trong sheet(key: tên cột tương ứng với value của cột đó)
                    row_data = dict(zip(self.Tab_Col_Excel[k], row))
                    tuple1=list()
                    # Với các cột SQL nếu có trong các cột của excel thì lấy giá trị tương ứng(Lấy giá trị theo hàng)
                    for column in self.Tab_Col_SQL[k]:
                        if column in row_data:
                            tuple1.append(row_data[column])
                        else:
                            tuple1.append(None)
                    # Add list dưới dạng tuple
                    list1.append(tuple(tuple1))

                # Check trùng, none . List1: Insert, List2: Update
                list1,list2 = self.check_ID(list1,k)

                # Add to dict
                res[k]=list1
                res1[k]=list2

        return res,res1
    def __Tab_Col_Excel__(self,wb):
        #Lấy Bảng và Cột Excel

        # Dict{} với key: tên sheet, values: các cột trong sheet
        res = {}
        sheet_names = wb.sheetnames
        for sheet in sheet_names:
            sht=wb[sheet]
            column_names = [cell.value for cell in sht[2] ]
            res[sheet] = column_names
        return res
    def __Tab_Col_SQL__(self):
        #Lấy Bảng và Cột SQL

        # Data SQL
        res = {}

        # len table
        res1 = {}

        # name table in SQL
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = self.cursor.fetchall()

        #Tạo dict chứa keys: tên bảng, values: tên các thuộc tính của bảng
        for  table in tables:
            list2=[]

            # Lấy tên cột
            k = f'PRAGMA table_info({table[0]})'
            self.cursor.execute(k)
            columns = self.cursor.fetchall()

            # Tạo list chứa tên cột
            for column in columns:
                list2.append(column[1])

            # Add to dict
            # res {'BUS': ['NO', 'NAME', 'kV', 'Code'], 'Line': ['NO','RATE_A']}
            # res1 số lượng cột có trong table
            res[table[0]] = list2
            res1[table[0]] = len(columns)

        return res,res1
    def print(self):

        dicts = [{'__INSERT__':self.Insert}, {'__UPDATE__': self.Update}]

        for res in dicts:
            for key, value in res.items():
                print(f"{key}:")
                if len(value) == 0:
                    print("None")
                else:
                    for sub_key, sub_value in value.items():
                        print(sub_key)
                        if len(sub_value) == 0:
                            print("None")
                        else:
                            for item in sub_value:
                                print(item)
                print()
                print()
    def check_ID(self,list1,k):

        ## Delete Tuple None
        list1 = [t for t in list1 if t[0] is not None]

        ## Check ID trùng
        sql=f'SELECT {self.Tab_Col_SQL[k][0]} FROM {k} '
        self.cursor.execute(sql)
        rows = self.cursor.fetchall()

        ## Tạo List Update
        list2=[]
        for row in rows:
            for first_values in list1:
                if first_values[0] == row[0]:
                    list1.remove(first_values)
                    list2.append(first_values)

        return list1,list2
    def main(self):
        """Nhập dữ liệu vào SQL"""

        """Đối chiếu cột trong SQL với Excel"""
        for table in self.Tab_Col_SQL.keys():
            if table in self.Tab_Col_Excel.keys():

                """INSERT"""
                ## Tạo câu lệnh Insert: INSERT INTO BUS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                sql=f"INSERT INTO {table} VALUES ({', '.join(['?'] * self.Len_Tab[table])})"
                self.cursor.executemany(sql,self.Insert[table])

                """UPDATE"""
                ## (List3:tạo 1 list tuple chứa cả data update, add ID vào cuối tuple)
                list3 = [t + (t[0],) for t in self.Update[table]]

                ## Tạo câu lệnh Update
                ## set_sql:   NO = ?, NAME = ?, kV = ?, FLAG = ?, PLOAD_kw = ?, QLOAD_kvar = ?, Qshunt_kvar = ?, Vscheduled = ?, Code = ?
                set_sql = ', '.join(f'{col} = ?' for col in self.Tab_Col_SQL[table])
                ## sql= UPDATE BUS SET NO = ?, NAME = ?, kV = ?, FLAG = ?, PLOAD_kw = ?, QLOAD_kvar = ?, Qshunt_kvar = ?, Vscheduled = ?, Code = ? Where NO=?
                sql=f'UPDATE {table} SET {set_sql} Where {self.Tab_Col_SQL[table][0]}=?'

                self.cursor.executemany(sql, list3)

        self.conn.commit()
        self.conn.close()
def taodatabase(name):
    #tạo bảng và các dữ liệu
    conn=sqlite3.connect(name)
    cursor=conn.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS  BUS(NO,NAME,kV,Code)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  LINE(NO,FROMBUS,TOBUS,CID,Rpu,Xpu,Gpu,Bpu,Length)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  GEN(NO,CID,Vsched,P,Q,Qmax,Qmin,Xd)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  LOAD(NO,CID,P,Q)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  OPTION(NAMEOPT,VALUE)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  BRANCH_RESULT(NO,FROMBUS,TOBUS,CID,P1,Q1,I1,P2,Q2,I2)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  BUS_RESULT(NO,MAG,ANG)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  X2TRANSFORMER(NO,FROMBUS,TOBUS,Rpu,Xpu,Gpu,Bpu,Winding1,Winding2,S_MVA)')

    cursor.execute('CREATE TABLE IF NOT EXISTS  SHUNT(NO,BUS,ID,Q_nom,U_nom)')
    conn.commit()
    conn.close()
    print("case created")
if __name__ == '__main__':
    name = '5nut.db'
    taodatabase(name)

    excel_file = '5nut.xlsx'
    db_file = '5nut.db'
    main = database(excel_file, db_file)
    main.print()
    main.main()
